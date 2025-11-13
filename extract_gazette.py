"""將立法院公報 XML 轉成 Excel 檔案的小工具。"""
import sys
import re
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import Workbook

def html_to_text(html: str) -> str:
    """把 HTMLContent 裡的 HTML 標籤粗略轉成純文字。"""
    if not html:
        return ""
    
    # 1) <br> / <br/> 換成換行
    html = re.sub(r"<br\s*/?>", "\n", html, flags=re.IGNORECASE)

    # 2) 移除所有 HTML 標籤，例如 <p>、<span>...
    text = re.sub(r"<[^>]+>", "", html)

    # 3) 處理一些常見 HTML 實體
    text = (
        text.replace("&nbsp;", " ")
            .replace("&amp;", "&")
            .replace("&lt;", "<")
            .replace("&gt;", ">")
    )

    # 4) 收斂空白與換行
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\s*\n\s*", "\n", text)
    return text.strip()

def get_text(record, tag_name: str) -> str:
    """從指定的 record 中取得某個子節點文字。"""
    el = record.find(tag_name)
    if el is not None and el.text:
        return el.text.strip()
    return ""

def extract_to_excel(xml_path: Path, output_path: Path) -> None:
    """從 XML 檔案中提取資料並寫入 Excel 檔案。"""
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # 找出所有recored
    records = root.findall("Record")
    print(f"找到 {len(records)} 筆Recored")

    # 我們要輸出的欄位（標題列）
    columns = [
        "MetaId",
        "Doc_Style_LName",
        "Doc_Style_SName",
        "Chapter",
        "PubGov",
        "PubGovName",
        "UndertakeGov",
        "Officer_name",
        "Date_Created",
        "Date_Published",
        "GazetteId",
        "Title",
        "ThemeSubject",
        "Keyword",
        "Explain",
        "Category",
        "Service",
        "GazetteHTML",
        "HTMLContentText",  # 特別欄位：處理過的純文字內文
    ]

    #建立新的excel
    wb = Workbook()

    ws_gazette = wb.active
    ws_gazette.title = "Gazette"

    # 寫入標題列
    for col_idx, name in enumerate(columns, start=1):
        ws_gazette.cell(row=1, column=col_idx, value=name)

    # Sheet 2: Lines（每一行文字一列）
    ws_lines = wb.create_sheet(title="Lines")
    lines_header = ["MetaId", "GazetteId", "LineNo", "Text"]
    for col_idx, name in enumerate(lines_header, start=1):
        ws_lines.cell(row=1, column=col_idx, value=name)

    # 目前寫到 Lines 的第幾列（從第 2 列開始）
    lines_row_idx = 2

    # Sheet 3: Articles（每條條文一列）
    ws_articles = wb.create_sheet(title="Articles")
    articles_header = ["MetaId", "GazetteId", "ArticleNo", "TitleLine",
                       "LineStart", "LineEnd", "Text"]

    for col_idx, name in enumerate(articles_header, start=1):
        ws_articles.cell(row=1, column=col_idx, value=name)
    articles_row_idx = 2

    # 「第X條」的條文標題判斷規則：支援中文數字與阿拉伯數字
    article_pattern = re.compile(
        r"^第\s*([一二三四五六七八九十零〇百千0-9]+)\s*條"
    )

    # === 寫入內容 ===
    for row_idx, record in enumerate(records, start=2):
        # 先準備 Gazette sheet 的資料
        row_values = {}

        # 先抓出 HTMLContent，因為兩個 sheet 都會用到
        raw_html = get_text(record, "HTMLContent")
        html_text = html_to_text(raw_html)

        for name in columns:
            if name == "HTMLContentText":
                row_values[name] = html_text
            else:
                row_values[name] = get_text(record, name)

        #把這筆record寫進excel的一列
        for col_idx, name in enumerate(columns, start=1):
            ws_gazette.cell(row=row_idx, column=col_idx, value=row_values[name])

        # 再寫入 Lines sheet：把 html_text 依照換行切成多段
        meta_id = row_values["MetaId"]
        gazette_id = row_values["GazetteId"]

        # 先把文字拆成「行」，同時寫進 Lines
        non_empty_lines = []  # [(line_no, text), ...]
        for line_no, line in enumerate(html_text.split("\n"), start=1):
            clean_line = line.strip()
            if not clean_line:
                continue  # 空行就跳過

            ws_lines.cell(row=lines_row_idx, column=1, value=meta_id)
            ws_lines.cell(row=lines_row_idx, column=2, value=gazette_id)
            ws_lines.cell(row=lines_row_idx, column=3, value=line_no)
            ws_lines.cell(row=lines_row_idx, column=4, value=clean_line)

            lines_row_idx += 1

            non_empty_lines.append((line_no, clean_line))

        # 再從 non_empty_lines 裡找「第X條」→ 組成 Articles
        current_article_no = None
        current_title_line = None
        current_start_line = None
        current_lines = []
        last_line_no = None

        for line_no, text in non_empty_lines:
            match = article_pattern.match(text)
            if match:
                # 遇到新的「第X條」，先把前一條寫進去
                if current_article_no is not None and current_lines:
                    article_text = "\n".join(current_lines).strip()
                    ws_articles.cell(row=articles_row_idx, column=1, value=meta_id)
                    ws_articles.cell(row=articles_row_idx, column=2, value=gazette_id)
                    ws_articles.cell(row=articles_row_idx, column=3, value=current_article_no)
                    ws_articles.cell(row=articles_row_idx, column=4, value=current_title_line)
                    ws_articles.cell(row=articles_row_idx, column=5, value=current_start_line)
                    ws_articles.cell(row=articles_row_idx, column=6, value=last_line_no)
                    ws_articles.cell(row=articles_row_idx, column=7, value=article_text)
                    articles_row_idx += 1

                # 開始新的條文
                current_article_no = match.group(1)
                current_title_line = text
                current_start_line = line_no
                current_lines = []
                last_line_no = line_no
            else:
                if current_article_no is not None:
                    current_lines.append(text)
                    last_line_no = line_no

        # 迴圈結束後，還有一條在累積，記得寫入
        if current_article_no is not None and current_lines:
            article_text = "\n".join(current_lines).strip()
            ws_articles.cell(row=articles_row_idx, column=1, value=meta_id)
            ws_articles.cell(row=articles_row_idx, column=2, value=gazette_id)
            ws_articles.cell(row=articles_row_idx, column=3, value=current_article_no)
            ws_articles.cell(row=articles_row_idx, column=4, value=current_title_line)
            ws_articles.cell(row=articles_row_idx, column=5, value=current_start_line)
            ws_articles.cell(row=articles_row_idx, column=6, value=last_line_no)
            ws_articles.cell(row=articles_row_idx, column=7, value=article_text)
            articles_row_idx += 1
        
    # 儲存Excel檔案
    wb.save(output_path)
    print(f"已將{len(records)}筆資料輸出到 {output_path}")
    print(f"Lines 工作表共寫入 {lines_row_idx - 2} 筆行文字")
    print(f"Articles 工作表共寫入 {articles_row_idx - 2} 筆條文")

def main():
    """命令列進入點：解析參數並呼叫 extract_to_excel。"""
    if len(sys.argv) !=3:
        print("用法: python extract_gazette.py <輸入XML檔案路徑> <輸出Excel檔案路徑>")
        print("範例: python extract_gazette.py 114-11-12/114-11-12.xml gazette_114-11-12.xlsx")
        sys.exit(1)
    
    xml_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    if not xml_path.exists():
        print(f"找不到輸入的 XML 檔案: {xml_path}")
        sys.exit(1)

    extract_to_excel(xml_path, output_path)

if __name__ == "__main__":
    main()