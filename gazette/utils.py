# gazette/utils.py
import re
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import Workbook
import pdfplumber
from typing import Optional 

def html_to_text(html: str) -> str:
    """把 HTMLContent 裡的 HTML 標籤粗略轉成純文字。"""
    if not html:
        return ""
    
    # 1) <br> / <br/> 換成換行
    html = re.sub(r"<br\s*/?>", "\n", html, flags=re.IGNORECASE)

    # 2) 移除所有 HTML 標籤，例如 <p>、<span>...
    text = re.sub(r"<[^>]+>", "", html)

    # 3) 處理一些常見 HTML 實體
    text = (
        text.replace("&nbsp;", " ")
            .replace("&amp;", "&")
            .replace("&lt;", "<")
            .replace("&gt;", ">")
    )

    # 4) 收斂空白與換行
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\s*\n\s*", "\n", text)
    return text.strip()

def get_text(record, tag_name: str) -> str:
    """從指定的 record 中取得某個子節點文字。"""
    el = record.find(tag_name)
    if el is not None and el.text:
        return el.text.strip()
    return ""

def extract_to_excel(xml_path: Path, output_path: Path) -> None:
    """從 XML 檔案中提取資料並寫入 Excel 檔案。"""
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # 找出所有recored
    records = root.findall("Record")
    print(f"找到 {len(records)} 筆Recored")

    # 我們要輸出的欄位（標題列）
    columns = [
        "MetaId",
        "Doc_Style_LName",
        "Doc_Style_SName",
        "Chapter",
        "PubGov",
        "PubGovName",
        "UndertakeGov",
        "Officer_name",
        "Date_Created",
        "Date_Published",
        "GazetteId",
        "Title",
        "ThemeSubject",
        "Keyword",
        "Explain",
        "Category",
        "Service",
        "GazetteHTML",
        "HTMLContentText",  # 特別欄位：處理過的純文字內文
    ]

    #建立新的excel
    wb = Workbook()

    ws_gazette = wb.active
    ws_gazette.title = "Gazette"

    # 寫入標題列
    for col_idx, name in enumerate(columns, start=1):
        ws_gazette.cell(row=1, column=col_idx, value=name)

    # Sheet 2: Lines（每一行文字一列）
    ws_lines = wb.create_sheet(title="Lines")
    lines_header = ["MetaId", "GazetteId", "LineNo", "Text"]
    for col_idx, name in enumerate(lines_header, start=1):
        ws_lines.cell(row=1, column=col_idx, value=name)

    # 目前寫到 Lines 的第幾列（從第 2 列開始）
    lines_row_idx = 2

    # Sheet 3: Articles（每條條文一列）
    ws_articles = wb.create_sheet(title="Articles")
    articles_header = ["MetaId", "GazetteId", "ArticleNo", "TitleLine",
                       "LineStart", "LineEnd", "Text"]

    for col_idx, name in enumerate(articles_header, start=1):
        ws_articles.cell(row=1, column=col_idx, value=name)
    articles_row_idx = 2

    # 「第X條」的條文標題判斷規則：支援中文數字與阿拉伯數字
    article_pattern = re.compile(
        r"^第\s*([一二三四五六七八九十零〇百千0-9]+)\s*條"
    )

    # === 寫入內容 ===
    for row_idx, record in enumerate(records, start=2):
        # 先準備 Gazette sheet 的資料
        row_values = {}

        # 先抓出 HTMLContent，因為兩個 sheet 都會用到
        raw_html = get_text(record, "HTMLContent")
        html_text = html_to_text(raw_html)

        for name in columns:
            if name == "HTMLContentText":
                row_values[name] = html_text
            else:
                row_values[name] = get_text(record, name)

        #把這筆record寫進excel的一列
        for col_idx, name in enumerate(columns, start=1):
            ws_gazette.cell(row=row_idx, column=col_idx, value=row_values[name])

        # 再寫入 Lines sheet：把 html_text 依照換行切成多段
        meta_id = row_values["MetaId"]
        gazette_id = row_values["GazetteId"]

        # 先把文字拆成「行」，同時寫進 Lines
        non_empty_lines = []  # [(line_no, text), ...]
        for line_no, line in enumerate(html_text.split("\n"), start=1):
            clean_line = line.strip()
            if not clean_line:
                continue  # 空行就跳過

            ws_lines.cell(row=lines_row_idx, column=1, value=meta_id)
            ws_lines.cell(row=lines_row_idx, column=2, value=gazette_id)
            ws_lines.cell(row=lines_row_idx, column=3, value=line_no)
            ws_lines.cell(row=lines_row_idx, column=4, value=clean_line)

            lines_row_idx += 1

            non_empty_lines.append((line_no, clean_line))

        # 再從 non_empty_lines 裡找「第X條」→ 組成 Articles
        current_article_no = None
        current_title_line = None
        current_start_line = None
        current_lines = []
        last_line_no = None

        for line_no, text in non_empty_lines:
            match = article_pattern.match(text)
            if match:
                # 遇到新的「第X條」，先把前一條寫進去
                if current_article_no is not None and current_lines:
                    article_text = "\n".join(current_lines).strip()
                    ws_articles.cell(row=articles_row_idx, column=1, value=meta_id)
                    ws_articles.cell(row=articles_row_idx, column=2, value=gazette_id)
                    ws_articles.cell(row=articles_row_idx, column=3, value=current_article_no)
                    ws_articles.cell(row=articles_row_idx, column=4, value=current_title_line)
                    ws_articles.cell(row=articles_row_idx, column=5, value=current_start_line)
                    ws_articles.cell(row=articles_row_idx, column=6, value=last_line_no)
                    ws_articles.cell(row=articles_row_idx, column=7, value=article_text)
                    articles_row_idx += 1

                # 開始新的條文
                current_article_no = match.group(1)
                current_title_line = text
                current_start_line = line_no
                current_lines = []
                last_line_no = line_no
            else:
                if current_article_no is not None:
                    current_lines.append(text)
                    last_line_no = line_no

        # 迴圈結束後，還有一條在累積，記得寫入
        if current_article_no is not None and current_lines:
            article_text = "\n".join(current_lines).strip()
            ws_articles.cell(row=articles_row_idx, column=1, value=meta_id)
            ws_articles.cell(row=articles_row_idx, column=2, value=gazette_id)
            ws_articles.cell(row=articles_row_idx, column=3, value=current_article_no)
            ws_articles.cell(row=articles_row_idx, column=4, value=current_title_line)
            ws_articles.cell(row=articles_row_idx, column=5, value=current_start_line)
            ws_articles.cell(row=articles_row_idx, column=6, value=last_line_no)
            ws_articles.cell(row=articles_row_idx, column=7, value=article_text)
            articles_row_idx += 1
        
    # 儲存Excel檔案
    wb.save(output_path)
    print(f"已將{len(records)}筆資料輸出到 {output_path}")
    print(f"Lines 工作表共寫入 {lines_row_idx - 2} 筆行文字")
    print(f"Articles 工作表共寫入 {articles_row_idx - 2} 筆條文")

def merge_chinese_lines(lines):
    """
    將多行文字合併成較符合中文閱讀的「段」。
    規則：
    - 空白行：結束當前段落。
    - 行尾遇到 。！？；：!?;: 之一：結束當前段落。
    """
    buffer = []
    end_punct = "。！？；：!?;:"

    for raw in lines:
        line = raw.strip()
        if not line:
            # 空行：直接結束一段
            if buffer:
                yield "".join(buffer).strip()
                buffer = []
            continue

        buffer.append(line)

        # 句尾有標點，就收斂成一段
        if line[-1] in end_punct:
            yield "".join(buffer).strip()
            buffer = []

    if buffer:
        yield "".join(buffer).strip()


def classify_segment(text: str, page: int, seg_no: int) -> str:
    """
    粗略判斷段落類型：
    - Header：第 1 頁前幾行
    - ArticleTitle：開頭包含「第X條」
    - Item：條列「一、」「二、」「3)」「(一)」...
    - Body：其他一般內文
    """
    t = text.strip()

    # 簡單把第 1 頁前幾行當作頁首
    if page == 1 and seg_no <= 3:
        return "Header"

    # 條文標題：第X條
    if re.match(r"^第\s*[一二三四五六七八九十零〇百千0-9]+\s*條", t):
        return "ArticleTitle"

    # 條列項目：一、 二、 三、
    if re.match(r"^[一二三四五六七八九十]+\s*、", t):
        return "Item"

    # 子項目：(一) 1) 2.
    if re.match(r"^[（(]?[一二三四五六七八九十0-9]+[)）．\.、]", t):
        return "SubItem"

    return "Body"


def extract_pdf_to_excel(
    pdf_path: Path,
    output_path: Path,
    original_name: Optional[str] = None,
) -> None:
    """
    從 PDF 檔案中抽取文字，寫入 Excel。
    - 每一頁拆成多個「中文段落」
    - 每段標記一個 Type（Header / ArticleTitle / Item / SubItem / Body）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PdfText"

    headers = ["SourceFile", "Page", "SegmentNo", "Type", "Text"]
    for col_idx, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=name)

    file_name = original_name or pdf_path.name
    row_idx = 2

    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            if not text.strip():
                continue

            # 先拿到原始行
            raw_lines = [ln for ln in (l.strip() for l in text.splitlines()) if ln]

            # 用中文規則合併成段
            segments = list(merge_chinese_lines(raw_lines))

            for seg_no, segment in enumerate(segments, start=1):
                seg_type = classify_segment(segment, page_number, seg_no)

                ws.cell(row=row_idx, column=1, value=file_name)
                ws.cell(row=row_idx, column=2, value=page_number)
                ws.cell(row=row_idx, column=3, value=seg_no)
                ws.cell(row=row_idx, column=4, value=seg_type)
                ws.cell(row=row_idx, column=5, value=segment)
                row_idx += 1

    wb.save(output_path)
    print(f"PDF {file_name} 已輸出 {row_idx - 2} 段到 {output_path}")
